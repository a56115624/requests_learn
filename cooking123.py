import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import time

#創建
wb = Workbook()
ws = wb.active
#讀取現有的檔案
# wb = load_workbook("recipe.xlsx")
# ws = wb.active
title = ["id", "stored", "choice", "name", "URL", "ingredient", "img"]
ws.append(title)
ingredient = input("請輸入你需要的食材:")
for i in range(1):
  r1 = requests.get(f'https://www.wecook123.com/page/{i}/?s={ingredient}',
          headers = {
          'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36',
          'cookie': '_ga=GA1.2.385644580.1640405202; _gid=GA1.2.2094292870.1640405202; _fbp=fb.1.1640405202566.530207548; _gat=1; _gat_gtag_UA_68516311_1=1; _gat_gtag_UA_194976822_1=1'
      })
  soup = BeautifulSoup(r1.text,'html.parser')
  i = 1
  for b2 in soup.find_all('div',{'class':'fusion-post-content-wrapper'}):
    next_page = f"{b2.find('a').attrs['href']}" 
    recipe = []
    recipe.append(i)
    recipe.append("")
    recipe.append("")
    recipe.append(b2.find('h2',{'class':'entry-title fusion-post-title'}).text.strip())
    recipe.append(f"{b2.find('a').attrs['href']}")
  for b3 in soup.find_all('div',{'class':'fusion-image-wrapper'}):
    recipe.append(f"{b3.find('img').attrs['src']}") 
  # 取得食譜網站內頁
    r2 = requests.get(f"{next_page}",
          headers = {
          'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36',
          'cookie': '_ga=GA1.2.385644580.1640405202; _gid=GA1.2.2094292870.1640405202; _fbp=fb.1.1640405202566.530207548; _gat=1; _gat_gtag_UA_68516311_1=1; _gat_gtag_UA_194976822_1=1'
      })
    soup2 = BeautifulSoup(r2.text,'html.parser')
    for Material in soup2.find_all('tr',{'style':'height: 24px;'}):
      ingredient_other = Material.find('td').text.strip()
      recipe.append(",".join(ingredient_other))
ws.append(recipe)
