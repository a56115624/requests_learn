# -*- coding: utf-8 -*-
"""crawl_icook_and_kikkoman.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1_tg5-uasjyIAOW-dibX_qjqec-OC4_IR
"""

# -*- coding: utf-8 -*-
"""crawl_recipe.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1cs6gwy43B0vFpQNRAZ4GcojnzXYEpyDz
"""

# 龜甲萬
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

#創建
# wb = Workbook()
# ws = wb.active
#讀取現有的檔案
wb = load_workbook("recipe.xlsx")
ws = wb.active
# title = ["id", "stored", "choice", "name", "URL", "ingredient", "img"]
# ws.append(title)
ingredient = input("請輸入你需要的食材:")

r1 = requests.get("https://www.kikkoman.com.tw/search.html",
        params = {
            "q" : ingredient,
            "filter[q]" : "1"
        }
    )

soup = BeautifulSoup(r1.text, 'html.parser')

i = 0
for r in soup.find_all('div',{'class':'col-4 mb-5'}):
  rID = r.find('a').attrs['href'].split('-')[1].split('.')[0]
  rName = r.find('h3').text
  recipe = []
  i += 1
  recipe.append(i)
  recipe.append("")
  recipe.append("")
  recipe.append(rName)
  recipe.append(f"https://www.kikkoman.com.tw/{r.find('a').attrs['href']}")
  # 取得食譜網站內頁
  r2 = requests.get("https://www.kikkoman.com.tw/cookbook.php",
        params = {
            "rName" : rName,
            "rID" : rID
        }
    )
  soup2 = BeautifulSoup(r2.text,'html.parser')
  # 取得所有食材
  ingredients = []
  for r2 in soup2.find_all('div',{'class':'col-3'}):
      for r3 in r2.find_all('li'):
        for ingredient_name in range(len(r3.text)):
          try:
            if type(int(r3.text[ingredient_name])) == type(0):
              ingredients.append(r3.text.split(r3.text[ingredient_name])[0])
            break
          except:
            pass
  for img in soup2.find_all('div',{'id':'photo'}):
    img_src = img.find('img').attrs['src'].split('.')
    img_url = "https://www.kikkoman.com.tw" + img_src[1] + ".jpg" 
  # 取得食材
  ingredient_other = []
  for j in range(len(ingredients)):
    ingredient_other.append(ingredients[j])
  recipe.append(",".join(ingredient_other))
  recipe.append(img_url)
  ws.append(recipe)

wb.save("recipe.xlsx")