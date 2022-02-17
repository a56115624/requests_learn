import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

#創建
# wb = Workbook()
# ws = wb.active
#讀取現有的檔案
wb = load_workbook("recipe.xlsx")
ws = wb.active
# title = ["id", "stored", "choice", "name", "URL", "ingredient", "img"]
# ws.append(title)
ingredient = input("請輸入你需要的食材:")

r1 = requests.get(f'https://icook.tw/search/{ingredient}/?page=1',
      headers = {
      'user-agent' : 'Mozilla/5.0 (Windows NT 6.2; Win64; x64; rv:94.0) Gecko/20100101 Firefox/94.0',
      'cookie' : 'visitor=13729120841602549317; dable_uid=86130283.1633677780858; _pbjs_userid_consent_data=3524755945110770; _ga=GA1.2.1930250049.1636638806; _gcl_au=1.1.6672339.1636638806; _ga_Q65WJCEHK3=GS1.1.1637675129.2.1.1637675494.60; __auc=f835a1af17d0f46563e83aaf332; _fbp=fb.1.1636638808126.357671361; __gads=ID=4f73dca6be358f80-22cfae79a2ce00b2:T=1636638811:S=ALNI_MYKh_x0rmhnTitV-PGLjAOqnkKXUQ; _cc_id=898a48e4705061c5547518efe36156af; _ga_ZKZX6M179R=GS1.1.1637675181.2.0.1637675181.60; CF-IPCountry=TW; CSRF-TOKEN=r8139%2Bw15%2Bcd3rooBM1H1ECc6p1PvwHzP4aEEO5fyqG85OrKAJJtnMkClsXJpH5c6Yuprx1FeGrzbxz%2B8dcf9A%3D%3D; _icook_sess=UDNQUkpuU1c5WDY0VTVGQlBJeks2V3AxanNFOFNUb0ZxWS8vcUlURmN4Mk9PRHdreGRGRWFoL1Z6RzBRajVDbytHb2ZuZHhiQ2RtakhrZWFvc1Y4K1VZejdnQ1pBWnpKNWRtdHNiWFBsOEYycEYyTnh1R01pTDl1V0JoTjk1L3p5eUtoN2JvRFJCdjRSbXdWQ3YwQ1ZLQ1htVUgzMmNZM0haVFZncTBadjFQSzdjbTFxNlBGWTcrZGh2NjBqOE00aXV3ZEI2RmlUWjJta1F4MlJlN3RCZnBrbDFWWFZ1citKZGtIWXE3dlhnSmE3RjRFS2ZZUGo2N0ZtZFptTlhIZy0teGhOZ1IwVlB1L2hURDByMTFkVWFLZz09--fcb194f3ecfae2a9c96dbdddd3322af029c660c8; __asc=6702a64717d4d0b74ed8f1b286f; AMP_TOKEN=%24NOT_FOUND; _gid=GA1.2.953040650.1637675184; sent-cid=1637675193'
  })
soup = BeautifulSoup(r1.text,'html.parser')

i = 1
for r in soup.find_all('li',{'class':'browse-recipe-item'}):
  ingredients = r.find('p',{'class':'browse-recipe-content-ingredient'}).text.split('：')[1]
  recipe = []
  recipe.append(i)
  recipe.append("")
  recipe.append("")
  recipe.append(r.find('h2',{'class':'browse-recipe-name'}).text.strip())
  recipe.append(f"https://icook.tw{r.find('a').attrs['href']}")
  igredient_other = []
  for j in range(len(ingredients.split('、'))):
    igredient_other.append(ingredients.split('、')[j])
  igredient_other[-1] = igredient_other[-1].strip()
  recipe.append(",".join(igredient_other))
  i += 1

  # 取得食譜圖片
  recipe.append(r.find('img').attrs['data-src']) 
  ws.append(recipe)

wb.save("recipe.xlsx")
